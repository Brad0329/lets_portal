import sqlite3
import os
import bcrypt
from config import DB_PATH


def get_connection():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())


def init_db():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'staff',
            must_change_pw INTEGER DEFAULT 1,
            perm_bid_tag INTEGER DEFAULT 0,
            perm_display INTEGER DEFAULT 0,
            perm_keyword INTEGER DEFAULT 0,
            perm_org INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            expires_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS notice_tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            notice_id INTEGER NOT NULL,
            tag TEXT NOT NULL,
            tagged_by INTEGER,
            memo TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            UNIQUE(notice_id),
            FOREIGN KEY (notice_id) REFERENCES bid_notices(id),
            FOREIGN KEY (tagged_by) REFERENCES users(id)
        );
    """)

    # 초기 관리자 계정 (admin / admin1234)
    cursor.execute("SELECT id FROM users WHERE username='admin'")
    if not cursor.fetchone():
        cursor.execute(
            "INSERT INTO users (username, name, password_hash, role, must_change_pw) VALUES (?, ?, ?, ?, ?)",
            ("admin", "관리자", hash_password("admin1234"), "admin", 0),
        )

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS bid_notices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            title TEXT NOT NULL,
            organization TEXT,
            category TEXT,
            bid_no TEXT,
            start_date TEXT,
            end_date TEXT,
            status TEXT,
            url TEXT,
            keywords TEXT,
            collected_at TEXT,
            UNIQUE(source, bid_no)
        );

        CREATE TABLE IF NOT EXISTS organizations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT,
            page_category TEXT,
            url TEXT NOT NULL,
            is_active INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS keywords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            keyword TEXT NOT NULL,
            keyword_group TEXT,
            is_active INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS display_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            setting_key TEXT NOT NULL UNIQUE,
            setting_value TEXT NOT NULL,
            description TEXT,
            updated_at TEXT
        );
    """)

    # Phase 3: bid_notices 스키마 확장 (9개 컬럼 추가)
    new_columns = [
        ("file_url", "TEXT"),       # 첨부파일 URL (중기부 fileUrl)
        ("detail_url", "TEXT"),     # 상세페이지 URL (K-Startup detl_pg_url)
        ("apply_url", "TEXT"),      # 신청 URL (K-Startup biz_aply_url)
        ("region", "TEXT"),         # 지원 지역 (K-Startup supt_regin)
        ("target", "TEXT"),         # 지원 대상 (K-Startup aply_trgt)
        ("content", "TEXT"),        # 공고 내용 요약
        ("budget", "TEXT"),         # 지원 규모/예산 (중기부 supt_scale)
        ("contact", "TEXT"),        # 문의처 (K-Startup prch_cnpl_no)
        ("est_price", "TEXT"),      # 추정 가격 (나라장터 presmptPrce)
        ("apply_method", "TEXT"),   # 접수 방법 (K-Startup 온라인/방문 등)
        ("biz_enyy", "TEXT"),       # 창업 연차 조건 (K-Startup)
        ("target_age", "TEXT"),     # 대상 연령 (K-Startup)
        ("department", "TEXT"),     # 담당부서 (K-Startup biz_prch_dprt_nm)
        ("excl_target", "TEXT"),    # 제외 대상 (K-Startup aply_excl_trgt_ctnt)
        ("biz_name", "TEXT"),       # 통합공고 사업명 (K-Startup intg_pbanc_biz_nm)
        # 나라장터 전용 확장 필드
        ("bid_method", "TEXT"),      # 입찰 방식 (bidMethdNm)
        ("contract_method", "TEXT"), # 계약 체결 방법 (cntrctCnclsMthdNm)
        ("award_method", "TEXT"),    # 낙찰 방식 (sucsfbidMthdNm)
        ("open_date", "TEXT"),       # 개찰 일시 (opengDt)
        ("assign_budget", "TEXT"),   # 배정 예산 (asignBdgtAmt)
        ("contact_name", "TEXT"),    # 담당자 (ntceInsttOfclNm)
        ("contact_phone", "TEXT"),   # 담당자 전화 (ntceInsttOfclTelNo)
        ("contact_email", "TEXT"),   # 담당자 이메일 (ntceInsttOfclEmailAdrs)
        ("tech_eval_ratio", "TEXT"), # 기술평가 비율 (techAbltEvlRt)
        ("price_eval_ratio", "TEXT"),# 가격평가 비율 (bidPrceEvlRt)
        ("procure_class", "TEXT"),   # 조달 분류 (pubPrcrmntLrgClsfcNm > MidClsfcNm)
        ("attachments", "TEXT"),     # 첨부파일 JSON [{name, url}, ...]
    ]
    for col_name, col_type in new_columns:
        try:
            cursor.execute(f"ALTER TABLE bid_notices ADD COLUMN {col_name} {col_type}")
        except Exception:
            pass  # 이미 존재하면 무시

    # 설정 초기값 삽입 (없을 때만)
    defaults = [
        ("status_filter", "ongoing", "진행중만=ongoing, 전체=all"),
        ("date_range_days", "30", "입찰공고 최근 N일 이내 표시"),
        ("sort_order", "deadline", "마감임박순=deadline, 입찰공고일 최신순=latest"),
        ("items_per_page", "20", "페이지당 표시 건수"),
    ]
    for key, value, desc in defaults:
        cursor.execute(
            "INSERT OR IGNORE INTO display_settings (setting_key, setting_value, description, updated_at) VALUES (?, ?, ?, datetime('now'))",
            (key, value, desc),
        )

    # Phase A.2: collect_sources 테이블
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS collect_sources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            collector_type TEXT NOT NULL,
            is_active INTEGER DEFAULT 1,
            last_collected_at TEXT,
            last_collected_count INTEGER,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)

    # collect_sources 초기 데이터 (없을 때만)
    initial_sources = [
        ("나라장터", "nara"),
        ("K-Startup", "kstartup"),
        ("중소벤처기업부", "mss_biz"),
        ("창조경제혁신센터", "ccei"),
        # Phase A.3: 개별 기관 스크래핑 (28개)
        ("강원관광재단", "scraper"),
        ("강원정보문화산업진흥원", "scraper"),
        ("경기도경제과학진흥원", "scraper"),
        ("경기도일자리재단", "scraper"),
        ("경남관광재단", "scraper"),
        ("경남문화예술진흥원", "scraper"),
        ("경남콘텐츠코리아랩", "scraper"),
        ("광주정보문화산업진흥원", "scraper"),
        ("국토연구원", "scraper"),
        ("서울테크노파크", "scraper"),
        ("신용보증기금", "scraper"),
        ("원광대학교", "scraper"),
        ("전남정보문화산업진흥원", "scraper"),
        ("전라북도경제통상진흥원", "scraper"),
        ("전라북도콘텐츠융합진흥원", "scraper"),
        ("정보통신산업진흥원", "scraper"),
        ("제주관광공사", "scraper"),
        ("창업진흥원", "scraper"),
        ("충남테크노파크", "scraper"),
        ("충청북도과학기술혁신원", "scraper"),
        ("충청북도기업진흥원", "scraper"),
        ("포항테크노파크", "scraper"),
        ("한국관광공사", "scraper"),
        ("한국디자인진흥원", "scraper"),
        ("한국보육진흥원", "scraper"),
        ("한국산업기술진흥원", "scraper"),
        ("한국콘텐츠진흥원", "scraper"),
        ("한국환경산업기술원", "scraper"),
        # Phase A.3 추가분
        ("경기대진테크노파크", "scraper"),
        ("세종테크노파크", "scraper"),
        ("한국발명진흥회", "scraper"),
        ("인천테크노파크", "scraper"),
        ("건국대학교", "scraper"),
        ("대전기업정보포털", "scraper"),
        ("경남테크노파크", "scraper"),
        ("제주콘텐츠진흥원", "scraper"),
        ("대전정보문화산업진흥원", "scraper"),
        ("전주정보문화산업진흥원", "scraper"),
        ("한국예탁결제원", "scraper"),
        ("한국지식재산보호원", "scraper"),
        ("부산창업포탈", "scraper"),
        # CCEI 입찰공고 7개 지역
        ("CCEI-경기", "scraper"),
        ("CCEI-경남", "scraper"),
        ("CCEI-대구", "scraper"),
        ("CCEI-부산", "scraper"),
        ("CCEI-세종", "scraper"),
        ("CCEI-인천", "scraper"),
        ("CCEI-충북", "scraper"),
    ]
    for name, ctype in initial_sources:
        cursor.execute(
            "INSERT OR IGNORE INTO collect_sources (name, collector_type) VALUES (?, ?)",
            (name, ctype),
        )

    # 나라장터 관심 중분류 테이블
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS nara_interest_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            large_class TEXT NOT NULL,
            mid_class TEXT NOT NULL,
            is_active INTEGER DEFAULT 1,
            UNIQUE(large_class, mid_class)
        )
    """)
    # 초기 데이터 (엑셀 기반)
    interest_categories = [
        ("연구조사서비스", "학술연구서비스"),
        ("연구조사서비스", "시장 및 여론조사"),
        ("교육 및 전문직종/기술서비스", "교육서비스"),
        ("교육 및 전문직종/기술서비스", "환경관리 서비스"),
        ("교육 및 전문직종/기술서비스", "기술시험, 검사 및 분석"),
        ("교육 및 전문직종/기술서비스", "문화재 조사/발굴 및 수리"),
        ("교육 및 전문직종/기술서비스", "보건서비스"),
        ("교육 및 전문직종/기술서비스", "회계서비스"),
        ("교육 및 전문직종/기술서비스", "법무서비스"),
        ("교육 및 전문직종/기술서비스", "번역 및 통역서비스"),
        ("행사관리 및 기타 사업 지원서비스", "기타"),
        ("행사관리 및 기타 사업 지원서비스", "행사 기획 및 대행"),
        ("행사관리 및 기타 사업 지원서비스", "전시관 및 홍보관 설치"),
        ("행사관리 및 기타 사업 지원서비스", "농.림.어업 서비스"),
        ("여행, 숙박, 음식, 운송 및 보관서비스", "여행서비스"),
        ("매체제작, 디자인, 홍보/마케팅 서비스", "매체제작"),
        ("매체제작, 디자인, 홍보/마케팅 서비스", "홍보 및 마케팅"),
    ]
    for lg, mid in interest_categories:
        cursor.execute(
            "INSERT OR IGNORE INTO nara_interest_categories (large_class, mid_class) VALUES (?, ?)",
            (lg, mid),
        )

    # keywords 테이블에 source_id 컬럼 추가 (NULL=공통, 값=출처 전용)
    try:
        cursor.execute("ALTER TABLE keywords ADD COLUMN source_id INTEGER REFERENCES collect_sources(id)")
    except Exception:
        pass  # 이미 존재하면 무시

    # Phase C: 회사 프로필 테이블 (유연한 key-value 구조)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS company_profile (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            field_key TEXT NOT NULL UNIQUE,
            field_label TEXT NOT NULL,
            field_value TEXT DEFAULT '',
            field_type TEXT DEFAULT 'text',
            sort_order INTEGER DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # 회사 프로필 초기 항목
    profile_fields = [
        # 기본정보
        ("기본정보", "company_name", "회사명", "", "text", 1),
        ("기본정보", "representative", "대표자", "", "text", 2),
        ("기본정보", "established", "설립연도", "", "number", 3),
        ("기본정보", "industry", "업종", "", "text", 4),
        ("기본정보", "location", "소재지", "", "text", 5),
        ("기본정보", "is_sme", "중소기업 해당", "true", "boolean", 6),
        # 사업실적
        ("사업실적", "project_history", "수행실적", "[]", "json", 10),
        # 인력현황
        ("인력현황", "staff_list", "투입인력", "[]", "json", 20),
        # 재무상태
        ("재무상태", "annual_revenue", "연매출", "", "text", 30),
        ("재무상태", "credit_rating", "신용등급", "", "text", 31),
        ("재무상태", "capital", "자본금", "", "text", 32),
        # 기술력
        ("기술력", "technologies", "보유기술/방법론", "", "textarea", 40),
        ("기술력", "patents_certs", "특허/인증", "", "textarea", 41),
        # 참여제한
        ("참여제한", "region_available", "지역제한 가능 여부", "", "text", 50),
        ("참여제한", "consortium_available", "컨소시엄 가능 여부", "", "text", 51),
        ("참여제한", "subcontract_limit", "하도급 제한 사항", "", "text", 52),
    ]
    for cat, key, label, value, ftype, order in profile_fields:
        cursor.execute(
            "INSERT OR IGNORE INTO company_profile (category, field_key, field_label, field_value, field_type, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
            (cat, key, label, value, ftype, order),
        )

    # Phase C: AI 분석 결과 테이블
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ai_analyses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            notice_id INTEGER NOT NULL REFERENCES bid_notices(id),
            analysis_seq INTEGER NOT NULL DEFAULT 1,
            parsed_criteria TEXT,
            parsed_requirements TEXT,
            parsed_summary TEXT,
            total_score REAL,
            total_max_score REAL,
            match_rate REAL,
            match_detail TEXT,
            match_recommendation TEXT,
            simple_grade TEXT,
            simple_reason TEXT,
            mode TEXT DEFAULT 'full',
            profile_snapshot TEXT,
            profile_changes TEXT,
            model_parsing TEXT,
            model_matching TEXT,
            cost_krw REAL DEFAULT 0,
            tokens_total INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(notice_id, analysis_seq)
        )
    """)

    # Phase C: AI 크레딧 테이블
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ai_credits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('charge', 'usage')),
            amount REAL NOT NULL,
            balance REAL NOT NULL,
            description TEXT,
            notice_id INTEGER,
            analysis_id INTEGER,
            model_used TEXT,
            tokens_input INTEGER,
            tokens_output INTEGER,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # Phase C: AI 설정 테이블
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ai_config (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            description TEXT,
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    ai_config_defaults = [
        ("model_parsing", "claude-haiku-4-5-20251001", "문서 파싱 모델"),
        ("model_matching", "claude-sonnet-4-20250514", "매칭 분석 모델"),
        ("exchange_rate", "1400", "환율 (USD → KRW)"),
    ]
    for key, value, desc in ai_config_defaults:
        cursor.execute(
            "INSERT OR IGNORE INTO ai_config (key, value, description) VALUES (?, ?, ?)",
            (key, value, desc),
        )

    # Phase C: bid_notices에 AI 상태 컬럼 추가
    ai_columns = [
        ("ai_status", "TEXT"),
        ("ai_match_rate", "REAL"),
        ("ai_recommendation", "TEXT"),
        ("attachment_dir", "TEXT"),
    ]
    for col_name, col_type in ai_columns:
        try:
            cursor.execute(f"ALTER TABLE bid_notices ADD COLUMN {col_name} {col_type}")
        except Exception:
            pass

    # Phase C-5: 마스터 프로필 추출기 — 제안서 소스
    # extractor의 SourceMeta와 매핑 (src_id UNIQUE). 진행 상태/비용을 DB에 적재하여
    # 다른 세션에서도 상태 조회 가능. 파일 자체는 data/proposals_assets/{src_id}/에 저장.
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS proposal_sources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            src_id TEXT UNIQUE NOT NULL,
            file_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            proposal_title TEXT,
            organization TEXT,
            year INTEGER,
            awarded INTEGER,
            status TEXT DEFAULT 'pending',
            progress REAL DEFAULT 0,
            total_sections INTEGER DEFAULT 0,
            processed_sections INTEGER DEFAULT 0,
            current_section TEXT,
            cost_krw REAL DEFAULT 0,
            input_tokens INTEGER DEFAULT 0,
            output_tokens INTEGER DEFAULT 0,
            model TEXT,
            error_message TEXT,
            uploaded_by INTEGER REFERENCES users(id),
            uploaded_at TEXT DEFAULT (datetime('now','localtime')),
            completed_at TEXT
        )
    """)

    # Phase C-5: 추출된 claim (6개 카테고리)
    # claim_id는 content-derived MD5 — 재실행/증분 업로드 시에도 안정적 매핑.
    # user_verified: 0=미검토, 1=승인, 2=거부/병합됨(UI에서 숨김).
    # merged_from JSON — 여러 claim을 병합한 경우 원본 claim_id 보존.
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS claims (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            claim_id TEXT UNIQUE NOT NULL,
            proposal_source_id INTEGER REFERENCES proposal_sources(id),
            category TEXT NOT NULL,
            statement TEXT NOT NULL,
            tags TEXT,
            proposal_sections TEXT,
            length_variants TEXT,
            evidence_refs TEXT,
            source_section TEXT,
            source_page INTEGER,
            source_start_line INTEGER,
            recurrence INTEGER DEFAULT 1,
            confidence REAL DEFAULT 0.8,
            user_verified INTEGER DEFAULT 0,
            merged_from TEXT,
            user_edited_statement TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # Phase C-5: 정량 힌트 (실적/인증/등록/특허/자격)
    # claims와 별개로 관리 — 사용자가 company_profile의 project_history / patents_certs에
    # 수동으로 편입하는 후보 리스트.
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS proposal_quantitative_hints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proposal_source_id INTEGER REFERENCES proposal_sources(id),
            type TEXT,
            name TEXT NOT NULL,
            year INTEGER,
            organization TEXT,
            section TEXT,
            extractor TEXT,
            user_action TEXT DEFAULT 'pending',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # 인덱스
    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ai_analyses_notice ON ai_analyses(notice_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ai_credits_type ON ai_credits(type)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_claims_category ON claims(category)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_claims_source ON claims(proposal_source_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_claims_verified ON claims(user_verified)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_proposal_sources_status ON proposal_sources(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_quant_hints_source ON proposal_quantitative_hints(proposal_source_id)")
    except Exception:
        pass

    conn.commit()
    conn.close()


def import_excel_data(excel_path):
    """엑셀에서 기관 목록 + 키워드를 DB에 임포트"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    conn = get_connection()
    cursor = conn.cursor()

    # Sheet1: 기관 목록 (2행부터 데이터, 1행은 헤더)
    ws1 = wb.worksheets[0]
    rows = list(ws1.iter_rows(min_row=3, values_only=True))  # 3행부터 (1행 제목, 2행 헤더)
    cursor.execute("DELETE FROM organizations")
    for row in rows:
        no, category, name, active, page_cat, url = row[:6]
        if not name or not url:
            continue
        is_active = 1 if active == "O" else 0
        cursor.execute(
            "INSERT INTO organizations (name, category, page_category, url, is_active) VALUES (?, ?, ?, ?, ?)",
            (str(name).strip(), str(category).strip() if category else None, str(page_cat).strip() if page_cat else None, str(url).strip(), is_active),
        )

    # Sheet2: 키워드 (2행부터 데이터)
    ws2 = wb.worksheets[1]
    header_row = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    groups = [str(h).strip() if h else f"그룹{i}" for i, h in enumerate(header_row[1:], 1)]

    cursor.execute("DELETE FROM keywords")
    for row in ws2.iter_rows(min_row=3, values_only=True):
        for i, keyword in enumerate(row[1:]):
            if keyword and str(keyword).strip():
                group_name = groups[i] if i < len(groups) else "기타"
                cursor.execute(
                    "INSERT INTO keywords (keyword, keyword_group, is_active) VALUES (?, ?, 1)",
                    (str(keyword).strip(), group_name),
                )

    conn.commit()
    conn.close()
    wb.close()


if __name__ == "__main__":
    import sys

    init_db()
    print("DB 초기화 완료")

    excel_path = os.path.join(os.path.dirname(__file__), "..", "●입찰 공고(사업 공고) 기관별 입찰 정리.xlsx")
    if os.path.exists(excel_path):
        import_excel_data(excel_path)
        print("엑셀 데이터 임포트 완료")

        # 확인
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM organizations")
        print(f"  기관: {cur.fetchone()[0]}건")
        cur.execute("SELECT COUNT(*) FROM keywords")
        print(f"  키워드: {cur.fetchone()[0]}건")
        cur.execute("SELECT COUNT(*) FROM display_settings")
        print(f"  설정: {cur.fetchone()[0]}건")
        conn.close()
    else:
        print(f"엑셀 파일 없음: {excel_path}")

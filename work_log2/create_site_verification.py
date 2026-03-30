"""site_verification.xlsx 재생성 스크립트 (2026-03-30 업데이트)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "사이트 검증 결과"

thin = Side(style='thin', color='000000')
border = Border(top=thin, bottom=thin, left=thin, right=thin)

hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor='1F4E79')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
title_font = Font(name='Arial', bold=True, size=14)

col_widths = {'A': 18, 'B': 6, 'C': 28, 'D': 80, 'E': 50}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

ws.merge_cells('A1:E1')
c = ws['A1']
c.value = "Phase A.3 스크래핑 대상 사이트 검증 결과 (2026-03-30 최종 업데이트)"
c.font = title_font
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

headers = ['구분', '#', '기관명', 'URL', '비고(사유)']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = border
ws.row_dimensions[2].height = 25

sections = [
    ("✅ 구현 완료 — HTML 스크래핑 (34개)", PatternFill('solid', fgColor='C6EFCE'), Font(name='Arial', bold=True, size=11, color='006100'), [
        (1, "강원관광재단", "https://www.gwto.or.kr/www/selectBbsNttList.do?bbsNo=3&key=23", "고시/공고"),
        (2, "강원정보문화산업진흥원", "https://www.gica.or.kr/Home/H40000/H40200/boardList", "입찰공고"),
        (3, "경기도경제과학진흥원", "https://www.gbsa.or.kr/board/bid_info.do", "입찰정보 전용"),
        (4, "경기도일자리재단", "https://www.gjf.or.kr/main/pst/list.do?pst_id=nara_market_bid", "입찰공고"),
        (5, "경남관광재단", "https://gnto.or.kr/sub04/sub04_01.php", "ul.board_ul 구조"),
        (6, "경남문화예술진흥원", "http://gcaf.or.kr/bbs/board.php?bo_table=sub3_7&me_code=a030", "Gnuboard"),
        (7, "경남콘텐츠코리아랩", "https://www.gnckl.or.kr/bbs/board.php?bo_table=notice3", "Gnuboard"),
        (8, "광주정보문화산업진흥원", "https://www.gicon.or.kr/board.es?mid=a10205000000&bid=0020", "board.es"),
        (9, "국토연구원", "https://www.krihs.re.kr/board.es?mid=a10602000000&bid=0012", "board.es"),
        (10, "서울테크노파크", "http://seoultp.or.kr/user/nd26539.do", "입찰공고"),
        (11, "신용보증기금", "https://alio.go.kr/occasional/bidList.do", "ALIO 공공기관 입찰"),
        (12, "원광대학교", "https://intra.wku.ac.kr/services/contract/cntrc.jsp", "계약정보 포털"),
        (13, "전남정보문화산업진흥원", "https://www.jcia.or.kr/cf/information/notice/tender/self.do", "입찰공고"),
        (14, "전라북도경제통상진흥원", "http://www.jbba.kr/bbs/board.php?bo_table=sub05_06_02", "Gnuboard (td.td_date)"),
        (15, "전라북도콘텐츠융합진흥원", "https://www.jcon.or.kr/board/list.php?bbsId=BBSMSTR_000000000003", "입찰공고"),
        (16, "정보통신산업진흥원", "https://www.nipa.kr/home/2-3", "조달청 연계"),
        (17, "제주관광공사", "https://ijto.or.kr/korean/Bd/list.php?btable=tender_info", "입찰정보"),
        (18, "창업진흥원", "https://www.kised.or.kr/board.es?mid=a10303000000&bid=0005", "board.es"),
        (19, "충남테크노파크", "https://www.ctp.or.kr/community/bid.do", "lxml parser 사용"),
        (20, "충청북도과학기술혁신원", "http://www.cbist.or.kr/home/sub.do?mncd=118", "입찰공고"),
        (21, "충청북도기업진흥원", "http://www.cba.ne.kr/home/sub.php?menukey=140&cate=00000010", "입찰공고"),
        (22, "포항테크노파크", "https://www.ptp.or.kr/main/board/index.do?menu_idx=114&manage_idx=3", "입찰정보"),
        (23, "한국관광공사", "https://touraz.kr/publicTenderList", "Touraz 입찰"),
        (24, "한국디자인진흥원", "http://www.kidp.or.kr/?menuno=1012", "내부입찰"),
        (25, "한국보육진흥원", "https://www.kcpi.or.kr/kcpi/cyberpr/tender.do", "입찰공고"),
        (26, "한국산업기술진흥원", "https://www.kiat.or.kr/front/board/boardContentsListPage.do?board_id=77", "입찰공고"),
        (27, "한국콘텐츠진흥원", "https://www.kocca.kr/kocca/tender/list.do?menuNo=204106&cate=01", "입찰정보"),
        (28, "한국환경산업기술원", "https://www.keiti.re.kr/site/keiti/ex/board/List.do?cbIdx=277", "ul.list.col5 구조"),
        (29, "경기대진테크노파크", "https://gdtp.or.kr/board/announcement", "div.colgroup.noti 구조"),
        (30, "세종테크노파크", "https://sjtp.or.kr/bbs/board.php?bo_table=notice01", "Gnuboard"),
        (31, "한국발명진흥회", "https://www.kipa.org/kipa/notice/kw_0402.jsp", "offset 기반 페이지네이션"),
        (32, "인천테크노파크", "https://www.itp.or.kr/intro.asp?tmid=14", "POST 페이지네이션 + 세션 쿠키"),
        (33, "건국대학교", "https://www.konkuk.ac.kr/konkuk/2243/subview.do", "K2Web CMS, SSO 통과 후 접근 가능"),
        (34, "대전기업정보포털", "https://www.dips.or.kr/pbanc?mid=a10201000000", "사업공고 (사실상 입찰), PDF 링크"),
    ]),
    ("✅ 구현 완료 — CCEI 입찰공고 JSON API (7개)", PatternFill('solid', fgColor='B7DEE8'), Font(name='Arial', bold=True, size=11, color='003366'), [
        (35, "CCEI-경기", "https://ccei.creativekorea.or.kr/gyeonggi/allim/allimList.json", "POST JSON API"),
        (36, "CCEI-경남", "https://ccei.creativekorea.or.kr/gyeongnam/allim/allimList.json", "POST JSON API"),
        (37, "CCEI-대구", "https://ccei.creativekorea.or.kr/daegu/allim/allimList.json", "POST JSON API"),
        (38, "CCEI-부산", "https://ccei.creativekorea.or.kr/busan/allim/allimList.json", "POST JSON API"),
        (39, "CCEI-세종", "https://ccei.creativekorea.or.kr/sejong/allim/allimList.json", "POST JSON API"),
        (40, "CCEI-인천", "https://ccei.creativekorea.or.kr/incheon/allim/allimList.json", "POST JSON API"),
        (41, "CCEI-충북", "https://ccei.creativekorea.or.kr/chungbuk/allim/allimList.json", "POST JSON API"),
    ]),
    ("🔧 추가 조치 필요 (4개)", PatternFill('solid', fgColor='FFEB9C'), Font(name='Arial', bold=True, size=11, color='9C6500'), [
        (42, "경남테크노파크", "https://www.gntp.or.kr", "URL변경 필요 (원래 more.co.kr 서버 다운)"),
        (43, "대전정보문화산업진흥원", "https://www.dicia.or.kr/sub.do?menuIdx=MENU_000000000000100", "SSL 인증서 오류 — verify=False 필요"),
        (44, "전주정보문화산업진흥원", "https://www.jica.or.kr/2025/inner.php?sMenu=A4000", "SSL 오류 + URL /2016/→/2025/ 변경"),
        (45, "한국예탁결제원", "https://www.ksd.or.kr/ko/about-ksd/ksd-news/bid-notice", "React SPA — headless browser 필요"),
    ]),
    ("🔄 URL 변경 필요 (2개)", PatternFill('solid', fgColor='FCD5B4'), Font(name='Arial', bold=True, size=11, color='974706'), [
        (46, "제주콘텐츠진흥원", "https://ofjeju.kr/communication/account/bid.htm", "기존 URL은 공지사항 → 입찰 URL로 변경"),
        (47, "소상공인시장진흥공단", "https://semas.or.kr/web/board/webBoardList.kmdc?bCd=220&pNm=BOA0102", "입찰정보 5건뿐, 실효성 낮음"),
    ]),
    ("🔧 JS 렌더링 보류 (1개)", PatternFill('solid', fgColor='E2EFDA'), Font(name='Arial', bold=True, size=11, color='375623'), [
        (48, "한국지식재산보호원", "https://www.koipa.re.kr/home/board/brdList.do?menu_cd=000042", "JS 렌더링 — BeautifulSoup 불가, API 또는 Selenium 필요"),
    ]),
    ("❌ 제외 (1개)", PatternFill('solid', fgColor='FFC7CE'), Font(name='Arial', bold=True, size=11, color='9C0006'), [
        (49, "부산창업포탈", "https://busanstartup.kr/biz_sup?mcode=biz02&deleteYn=N&busi_code=820", "창업지원 프로그램, 입찰 아님"),
    ]),
]

row = 3
data_font = Font(name='Arial', size=10)
url_font = Font(name='Arial', size=9, color='0563C1')

for section_label, section_fill, section_label_font, items in sections:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=section_label)
    c.font = section_label_font
    c.fill = section_fill
    c.alignment = Alignment(vertical='center')
    c.border = border
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = section_fill
        ws.cell(row=row, column=col).border = border
    ws.row_dimensions[row].height = 28
    row += 1

    for num, name, url, note in items:
        ws.cell(row=row, column=1, value="").border = border
        c2 = ws.cell(row=row, column=2, value=num)
        c2.font = data_font
        c2.alignment = Alignment(horizontal='center')
        c2.border = border
        c3 = ws.cell(row=row, column=3, value=name)
        c3.font = data_font
        c3.border = border
        c4 = ws.cell(row=row, column=4, value=url)
        c4.font = url_font
        c4.hyperlink = url
        c4.border = border
        c5 = ws.cell(row=row, column=5, value=note)
        c5.font = data_font
        c5.border = border
        ws.row_dimensions[row].height = 22
        row += 1

row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws.cell(row=row, column=1, value="※ 구현 완료 41개 사이트는 일괄 수집 버튼 1개로 수집 가능").font = Font(name='Arial', size=10, italic=True, color='666666')

out = r"C:\Users\user\Documents\lets_portal\work_log2\site_verification.xlsx"
wb.save(out)
print(f"Saved: {out}")
